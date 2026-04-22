import sys
import os
from datetime import datetime
import openpyxl
import calendar
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QMessageBox, QComboBox
)
from PyQt5.QtCore import Qt, QEvent
from PyQt5.QtGui import QFont
from PyQt5.QtGui import QPixmap

class MiniRupterApp(QMainWindow):
    def __init__(self):
        super().__init__()

        self.logo_path = r"J:\Production_Operations\Switches and Operators\Mini-Rupters\Resistance Test Log\Resistance Test Table\_internal\logo.png"
        self.save_dir = r"J:\Production_Operations\Switches and Operators\Mini-Rupters\Resistance Test Log\Resistance Test Table"
        
        self.file_path = os.path.join(self.save_dir, "Resistance_Test_Table.xlsx")
        self.last_values = {}
        self.fields = {}

        self.initUI()

    def initUI(self):
        self.setWindowTitle("Mini-Rupter Switch Resistance Test Log")
        self.setGeometry(100, 100, 900, 600)

        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)

        
        # Header
        header_widget = QWidget()
        header_widget.setStyleSheet("background-color: #47b8a4;")

        header_layout = QHBoxLayout(header_widget)
        header_layout.setContentsMargins(15, 10, 15, 10)
        header_layout.setSpacing(15)

        # Logo 
        logo_label = QLabel()
        if os.path.exists(self.logo_path):
            pixmap = QPixmap(self.logo_path)
            pixmap = pixmap.scaled(
                50, 50,  # smaller so it fits nicely
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation
            )
            logo_label.setPixmap(pixmap)

        logo_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        header_layout.addWidget(logo_label)

        # Title 
        title_label = QLabel("Mini-Rupter Switch Resistance Test Log")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setStyleSheet("color: white;")
        title_label.setAlignment(Qt.AlignCenter)

        header_layout.addWidget(title_label, stretch=1)

        main_layout.addWidget(header_widget)


        # Form layout
        form_layout = QVBoxLayout()
        form_layout.setSpacing(20)
        form_layout.setContentsMargins(100, 40, 100, 40)

        def create_row(label_text, widget):
            layout = QHBoxLayout()
            label = QLabel(label_text)
            label.setFont(QFont("Arial", 11, QFont.Bold))
            label.setMinimumWidth(180)

            widget.setMinimumHeight(40)
            widget.setFont(QFont("Arial", 11))

            layout.addWidget(label)
            layout.addWidget(widget)
            return layout

        # Date & Time
        self.datetime_input = QLineEdit()
        self.datetime_input.setReadOnly(True)
        form_layout.addLayout(create_row("Date and Time:", self.datetime_input))
        self.fields["Date and Time"] = self.datetime_input

        # Cat#
        self.cat_input = QLineEdit()
        form_layout.addLayout(create_row("Cat#:", self.cat_input))
        self.fields["Cat#"] = self.cat_input

        # JO#
        self.jo_input = QLineEdit()
        form_layout.addLayout(create_row("JO#:", self.jo_input))
        self.fields["JO#"] = self.jo_input

        # Operator ID
        self.operator_input = QLineEdit()
        form_layout.addLayout(create_row("Operator ID#:", self.operator_input))
        self.fields["Operator ID#"] = self.operator_input

        # Shift
        self.shift_combo = QComboBox()
        self.shift_combo.addItems(["", "AM", "PM"])
        form_layout.addLayout(create_row("Shift:", self.shift_combo))
        self.fields["Shift"] = self.shift_combo

        # Pad
        self.pad_combo = QComboBox()
        self.pad_combo.addItems(["", "Cu", "Al"])
        form_layout.addLayout(create_row("Pad:", self.pad_combo))
        self.fields["Pad"] = self.pad_combo

        # A∅ B∅ C∅
        self.a_input = QLineEdit()
        form_layout.addLayout(create_row("A∅:", self.a_input))
        self.fields["A∅"] = self.a_input

        self.b_input = QLineEdit()
        form_layout.addLayout(create_row("B∅:", self.b_input))
        self.fields["B∅"] = self.b_input

        self.c_input = QLineEdit()
        form_layout.addLayout(create_row("C∅:", self.c_input))
        self.fields["C∅"] = self.c_input

        # Submit button
        submit_btn = QPushButton("Submit")
        submit_btn.setMinimumHeight(50)
        submit_btn.setMinimumWidth(260)
        submit_btn.setFont(QFont("Arial", 12, QFont.Bold))
        submit_btn.setStyleSheet("""
            QPushButton {
                background-color: #47b8a4;
                color: white;
                border-radius: 5px;
            }
            QPushButton:hover { background-color: #3a9c8c; }
            QPushButton:pressed { background-color: #2d7d6d; }
        """)
        submit_btn.clicked.connect(self.submit_form)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_layout.addWidget(submit_btn)
        btn_layout.addStretch()

        form_layout.addSpacing(20)
        form_layout.addLayout(btn_layout)

        main_layout.addLayout(form_layout)
        main_widget.setLayout(main_layout)

        # Install event filters AFTER fields are created
        for widget in self.fields.values():
            widget.installEventFilter(self)

        self.update_datetime()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.FocusIn and obj in self.fields.values():
            field_name = next(k for k, v in self.fields.items() if v == obj)

            if field_name in self.last_values:
                last_val = self.last_values[field_name]

                # QLineEdit
                if isinstance(obj, QLineEdit):
                    if not obj.text().strip():
                        obj.setText(last_val)

                        # Delay selection so highlight is visible
                        QTimer.singleShot(0, obj.selectAll)

                # QComboBox
                elif isinstance(obj, QComboBox):
                    if not obj.currentText().strip():
                        index = obj.findText(last_val)
                        if index >= 0:
                            obj.setCurrentIndex(index)

        return super().eventFilter(obj, event)

    def update_datetime(self):
        self.datetime_input.setText(
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )

    def clear_form(self):
        # Clear fields except certain ones
        shift_index = self.shift_combo.currentIndex()
        pad_index = self.pad_combo.currentIndex()

        for key, widget in self.fields.items():
            if key not in ["Operator ID#", "Shift", "Pad", "Date and Time"]:
                widget.clear()

        self.shift_combo.setCurrentIndex(shift_index)
        self.pad_combo.setCurrentIndex(pad_index)

        self.update_datetime()

    def submit_form(self):
        cat = self.cat_input.text().strip()
        jo = self.jo_input.text().strip()
        operator = self.operator_input.text().strip()

        if not cat or not jo or not operator:
            QMessageBox.warning(self, "Error", "Cat#, JO#, and Operator ID# are required.")
            return

        data = [
            self.datetime_input.text(),
            cat,
            jo,
            operator,
            self.shift_combo.currentText(),
            self.pad_combo.currentText(),
            self.a_input.text().strip(),
            self.b_input.text().strip(),
            self.c_input.text().strip()
        ]

        self.last_values = {
            "Cat#": cat,
            "JO#": jo,
            "Operator ID#": operator,
            "Shift": self.shift_combo.currentText(),
            "Pad": self.pad_combo.currentText(),
            "A∅": self.a_input.text().strip(),
            "B∅": self.b_input.text().strip(),
            "C∅": self.c_input.text().strip()
        }

        try:
            self.write_to_excel(data)
            QMessageBox.information(self, "Success", "Record saved successfully.")
            self.clear_form()
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def write_to_excel(self, data):
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(
                f"Excel file not found:\n{self.file_path}"
            )

        wb = openpyxl.load_workbook(self.file_path)

        now = datetime.now()
        sheet_name = f"{now.strftime('%B')}_{now.year}"
        table_name = f"Table_{sheet_name}"

        headers = [
            "Date and Time", "Cat#", "JO#", "Operator ID#",
            "Shift", "Type of Pad", "A∅", "B∅", "C∅"
        ]

        # Create sheet if missing
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            
            column_widths = {
                "A": 19,  # Date and Time
                "B": 18,  # Cat#
                "C": 15,  # JO#
                "D": 14,  # Operator ID#
                "E": 9,  # Shift
                "F": 14,  # Type of Pad
                "G": 9,  # A∅
                "H": 9,  # B∅
                "I": 9,  # C∅
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            ws.row_dimensions[1].height = 30
            ws.row_dimensions[2].height = 30

            ws.merge_cells("G1:I1")
            ws["G1"] = "Value Recorded / µΩ"
            ws["G1"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )
            
            ws["G1"].fill = PatternFill(
                start_color="FFFF00",
                end_color="FFFF00",
                fill_type="solid"
            )

            ws["G1"].font = Font(bold=True)

            # Insert headers in row 2
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=2, column=col_num).value = header

            table = openpyxl.worksheet.table.Table(
                displayName=table_name,
                ref="A2:I2"
            )

            style = openpyxl.worksheet.table.TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )

            table.tableStyleInfo = style
            ws.add_table(table)
            center_align = openpyxl.styles.Alignment(
                horizontal="center", vertical="center"
            )
            for col in range(1, len(data) + 1):
                ws.cell(row=2, column=col).alignment = center_align
        else:
            ws = wb[sheet_name]

        # Append row
        ws.append(data)
        new_row = ws.max_row

        # Center-align newly added row
        center_align = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )
        for col in range(1, len(data) + 1):
            ws.cell(row=new_row, column=col).alignment = center_align

        # Extend the table safely
        table = ws.tables[table_name]
        start_cell, end_cell = table.ref.split(":")

        end_col_letter = "".join(filter(str.isalpha, end_cell))
        table.ref = f"A2:{end_col_letter}{new_row}"

        wb.save(self.file_path)
        wb.close()


def main():
    app = QApplication(sys.argv)
    window = MiniRupterApp()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()